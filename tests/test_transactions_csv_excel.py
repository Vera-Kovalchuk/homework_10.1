from src.transactions_csv_excel import read_csv, read_xlsx
import unittest
from unittest.mock import Mock, patch

import openpyxl


@patch("src.transactions_csv_excel.csv.DictReader")  # Подменяем pd.read_csv
def test_read_file_csv_success(mock_read_csv):
    mock_read_csv.return_value = Mock(to_dict=Mock(return_value=[]))
    result = read_csv("../files/transactions.csv")
    assert result == []


class TestTransactionLoaders(unittest.TestCase):

    def test_load_xlsx(self):
        mock_xlsx_data = [
            ("id", "state", "date", "amount", "currency_name", "currency_code", "from", "to", "description"),
            (1, "CANCELED", "2021-01-01", 100, "USD", 840, "Alice", "Bob", "Payment"),
            (2, "CANCELED", "2021-01-02", 200, "USD", 840, "Alice", "", "Transfer"),
            (4, "CANCELED", "2021-01-04", 300, "USD", 840, "", "Bob", "Invoice"),
        ]

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in mock_xlsx_data:
            sheet.append(row)

        with patch("openpyxl.load_workbook") as mock_load_workbook:
            mock_load_workbook.return_value = workbook
            transactions = read_xlsx("mock_file.xlsx")
            expected_transactions = [
                {
                    "id": 1,
                    "state": "CANCELED",
                    "date": "2021-01-01",
                    "operationAmount": {"amount": 100, "currency": {"name": "USD", "code": 840}},
                    "from": "Alice",
                    "to": "Bob",
                    "description": "Payment",
                },
                {
                    "id": 2,
                    "state": "CANCELED",
                    "date": "2021-01-02",
                    "operationAmount": {"amount": 200, "currency": {"name": "USD", "code": 840}},
                    "from": "Alice",
                    "to": "",  # Изменено с None на пустую строку
                    "description": "Transfer",
                },
                {
                    "id": 4,
                    "state": "CANCELED",
                    "date": "2021-01-04",
                    "operationAmount": {"amount": 300, "currency": {"name": "USD", "code": 840}},
                    "from": "",  # Изменено с None на пустую строку
                    "to": "Bob",
                    "description": "Invoice",
                },
            ]
            self.assertEqual(transactions, expected_transactions)
